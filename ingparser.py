#!/usr/bin/env python

#
# Copyright (c) 2015, 2017 Alexander Schrijver <alex@flupzor.nl>
#
# Permission to use, copy, modify, and distribute this software for any
# purpose with or without fee is hereby granted, provided that the above
# copyright notice and this permission notice appear in all copies.
#
# THE SOFTWARE IS PROVIDED "AS IS" AND THE AUTHOR DISCLAIMS ALL WARRANTIES
# WITH REGARD TO THIS SOFTWARE INCLUDING ALL IMPLIED WARRANTIES OF
# MERCHANTABILITY AND FITNESS. IN NO EVENT SHALL THE AUTHOR BE LIABLE FOR
# ANY SPECIAL, DIRECT, INDIRECT, OR CONSEQUENTIAL DAMAGES OR ANY DAMAGES
# WHATSOEVER RESULTING FROM LOSS OF USE, DATA OR PROFITS, WHETHER IN AN
# ACTION OF CONTRACT, NEGLIGENCE OR OTHER TORTIOUS ACTION, ARISING OUT OF
# OR IN CONNECTION WITH THE USE OR PERFORMANCE OF THIS SOFTWARE.
#

import csv
import sys
from datetime import datetime
from decimal import Decimal
from itertools import groupby
from operator import attrgetter
from optparse import OptionParser
from pprint import pprint

from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from transaction.transaction import INGTransaction


def main():
    usage = "{program_name}: file [file...]".format(program_name=sys.argv[0])
    parser = OptionParser(usage=usage)
    parser.add_option(
        '-d', action="store_true", dest='detailed', default=False,
        help='create a detailled overview, including the individual transactions')
    options, args = parser.parse_args()
    if len(args) == 0:
        parser.error("supply at least one file with transactions")

    path_list = args
    for path in path_list:
        with open(path, 'rb') as csvfile:
            parse_ing_file(csvfile, detailed=options.detailed)

def create_sheet(wb, month):
    ws = wb.create_sheet("{}".format(month))
    ws.column_dimensions[get_column_letter(1)].width = 22
    ws.column_dimensions[get_column_letter(2)].width = 15
    ws.column_dimensions[get_column_letter(3)].width = 42
    ws.column_dimensions[get_column_letter(4)].width = 22
    return ws

def add_category(ws, category, transactions):
    ws.append([("{}".format(category or 'geen')), ])

    total = Decimal('0')
    for transaction in transactions:
        total += transaction.amount
    ws.append([None, total])

def add_category_detailed(ws, category, transactions):
    ws.append([("{}".format(category or 'geen')), ])

    row_start = ws._current_row
    for transaction in transactions:
        ws.append([
            transaction.date, transaction.amount, transaction.description,
            transaction.contra_account])
    row_end = ws._current_row
    formula = '=SUM({column_letter}{row_start}:{column_letter}{row_end}'.format(
                column_letter=get_column_letter(2), row_start=row_start, row_end=row_end)
    ws.append([None, formula])

def parse_ing_file(file_handle, detailed=False):
    ing_reader = csv.reader(file_handle, delimiter=',')

    ing_reader.next()

    transactions = [INGTransaction.from_row(row) for row in ing_reader]
    transactions = sorted(transactions, key=attrgetter('date.year', 'date.month', 'category', 'date'))
    for year, month_group in groupby(transactions, attrgetter('date.year')):
        wb = Workbook()
        ws = wb.active
        wb.remove_sheet(ws)

        for month, category_group in groupby(month_group, attrgetter('date.month')):
            ws = create_sheet(wb, month)
            for category, group in groupby(category_group, attrgetter('category')):
                if detailed:
                    add_category_detailed(ws, category, group)
                else:
                    add_category(ws, category, group)
        wb.save("transactions-{}.xlsx".format(year))

if __name__ == '__main__':
    main()
